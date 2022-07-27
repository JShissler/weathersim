from copy import deepcopy
import random
import math
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule
from openpyxl.styles import Font, Border, Side, Alignment, Color, PatternFill

#### Classes
class BaseData:
    def __init__(self):
        pass

    def interpolate_data(self, data):
        year_data = []
        for month in range(len(data) - 1):
            month_data = []
            slope = data[month + 1] - data[month]
            for day in range(days_per_month):
                interpolated_day = day / days_per_month
                day_data = data[month] + (interpolated_day * slope)
                month_data.append(day_data)
            year_data.append(month_data)
        return year_data

    def generate_blank_list(self):
        data_full = []
        for year in range(calc_years):
            year_temp = []
            for month in range(months_per_year):
                month_temp = []
                for day in range(days_per_month):
                    day_temp = []
                    for hour in range(hours_per_day):
                        day_temp.append(None)
                    month_temp.append(day_temp)
                year_temp.append(month_temp)
            data_full.append(year_temp)
        return data_full

    def create_weighted_list_exponential(self, num_of_choices, peak_index, front_rate, back_rate = None):
        if not back_rate:
            back_rate = front_rate
        temp_odds = [0 for x in range(num_of_choices)]
        temp_odds[peak_index] = 10
        temp_odds_front = temp_odds[:peak_index + 1]
        temp_odds_back = temp_odds[peak_index:]
        for x in range(len(temp_odds_front) - 1):
            temp_odds_front[-(x+2)] = temp_odds_front[-(x+1)] / front_rate
        for x in range(len(temp_odds_back) - 1):
            temp_odds_back[x+1] = temp_odds_back[x] / back_rate
        temp_odds_front.pop()
        temp_odds = temp_odds_front + temp_odds_back
        return temp_odds

class TemperatureData(BaseData):
    def __init__(self, low, avg, high):
        super().__init__()
        self.low = self.interpolate_data(low)
        self.avg = self.interpolate_data(avg)
        self.high = self.interpolate_data(high)
        self.usable_hourly = self.generate_blank_list()

    def generate_temperatures(self, sunrise_param_data, sundown_param_data, cloud_param_data):
        print('Starting generate_temperatures')
        # Determine temperature based on sun and clouds as well as a random temperature fluctuation variable
        # Sun impact: Coldest at sunrise and temperature peak at 15:00 - Interpolation points of: Sunrise, Sunrise + 2 hours, 12:00, 15:00, sunset
        # Cloud impact: Clear and mostly clear increase temperature, partly clear has no impact, mostly cloudy and overcast reduce temperature
        # Temperature fluctuation: Can wander up and down, more likely to wander when close to zero, less likely when further away from zero - Calculated daily at sunset
        wandering_variation = 0
        wandering_correctiveness = 1 # Must be positive. A larger number means the temperature is more quickly corrected towards the average
        wandering_step_likelihood = 1.1 # Must be positive. A larger number means the temperature is less likely to wander by large amounts at once
        wandering_result_list = [x for x in range(-12, 13)]
        wandering_result_list = [x/2 for x in wandering_result_list]
        num_of_choices = len(wandering_result_list)

        # First, populate the usable_hours list with only the set temperature points
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    sunrise_hour = int(math.modf(sunrise_param_data[month][day])[1])
                    sunset_hour = int(math.modf(sundown_param_data[month][day])[1])
                    for hour in range(hours_per_day):
                        # Set the sunrise temperature point
                        if hour == sunrise_hour:
                            sunrise_temp = self.low[month][day] + wandering_variation
                            self.usable_hourly[year][month][day][hour] = sunrise_temp
                        # Set the 15:00 temperature point
                        elif hour == 15:
                            fifteen_temp = self.high[month][day] + wandering_variation
                            self.usable_hourly[year][month][day][hour] = fifteen_temp
                            # Set the Sunrise + 2 hours temperature point
                            sunrise_plus_two_temp = sunrise_temp + ((fifteen_temp - sunrise_temp) * 0.2)
                            self.usable_hourly[year][month][day][sunrise_hour + 2] = sunrise_plus_two_temp
                            # Set the 12:00 temperature point
                            twelve_temp = fifteen_temp - ((fifteen_temp - sunrise_temp) * 0.2)
                            self.usable_hourly[year][month][day][12] = twelve_temp
                        # Set the sunset temperature point
                        elif hour == sunset_hour:
                            sunset_temp = fifteen_temp - ((fifteen_temp - sunrise_temp) * 0.4)
                            self.usable_hourly[year][month][day][hour] = sunset_temp
                            # Calculate the new wandering variation
                            adjusted_wandering_variation = (wandering_variation / 100) * wandering_correctiveness
                            if wandering_variation >= 0:
                                wandering_odds = self.create_weighted_list_exponential(num_of_choices, math.floor(num_of_choices / 2), wandering_step_likelihood - adjusted_wandering_variation, wandering_step_likelihood)
                            else:
                                wandering_odds = self.create_weighted_list_exponential(num_of_choices, math.floor(num_of_choices / 2), wandering_step_likelihood, wandering_step_likelihood + adjusted_wandering_variation)
                            wandering_variation += random.choices(wandering_result_list, weights = wandering_odds, k = 1)[0]
        # Second, loop through the data again and interpolate temperature between the set temperatures found above
        beginning_interpolation_temp = 0
        ending_interpolation_temp = 0
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        distance_between_interpolation_temps = 0
                        # Check for a None datatype and then find the two temperatures it is between as well as the indexes between the temperatures
                        # Check if it is the first None of the usable_hours list, if so give a default beginning_interpolation_temp
                        if self.usable_hourly[year][month][day][hour] == None and day == 0 and month == 0 and year == 0 and hour < sunrise_param_data[0][0]:
                            first_day_index = 0
                            beginning_interpolation_temp = 32
                            for x in self.usable_hourly[year][month][day]:
                                if x != None:
                                    first_day_index = self.usable_hourly[year][month][day].index(x)
                                    break
                            ending_interpolation_temp = self.usable_hourly[year][month][day][first_day_index]
                            interpolated_temperatures = [(beginning_interpolation_temp + (ending_interpolation_temp - beginning_interpolation_temp) * x / first_day_index) for x in range(first_day_index)]
                            self.usable_hourly[year][month][day][:first_day_index] = interpolated_temperatures
                        # If an hour with a datatype of None is encountered, assign it the temperature immediately prior
                        elif self.usable_hourly[year][month][day][hour] == None:
                            changed_days = False
                            beginning_interpolation_temp = self.usable_hourly[year][month][day][hour - 1]
                            beginning_interpolation_temp_index = [year, month, day, hour - 1]
                            # Loop through an indefinite while loop to find the ending_interpolation_temp as well as the indexes between
                            loop_year = year
                            loop_month = month
                            loop_day = day
                            loop_hour = hour
                            while True:
                                # Check if loop_year is at the end of the calc_years. If so, give a default value for ending_interpolation_temp
                                if loop_year < calc_years:
                                    # Check if loop_month needs to roll over to the next year
                                    if loop_month < months_per_year:
                                        # Check if loop_day needs to roll over to the next month
                                        if loop_day < days_per_month:
                                            # Check if loop_hour needs to roll over to the next day
                                            if loop_hour < hours_per_day:
                                                if self.usable_hourly[loop_year][loop_month][loop_day][loop_hour] == None:
                                                    distance_between_interpolation_temps += 1
                                                    loop_hour += 1
                                                else:
                                                    ending_interpolation_temp = self.usable_hourly[loop_year][loop_month][loop_day][loop_hour]
                                                    ending_interpolation_temp_index = [loop_year, loop_month, loop_day, loop_hour]
                                                    break
                                            else:
                                                loop_hour = 0
                                                loop_day += 1
                                                changed_days = True
                                        else:
                                            loop_day = 0
                                            loop_month += 1
                                    else:
                                        loop_month = 0
                                        loop_year += 1
                                else:
                                    changed_days = False
                                    ending_interpolation_temp = self.usable_hourly[-1][-1][-2][6]
                                    ending_interpolation_temp_index = [-1, -1, -1, -1]
                                    break
                            # Calculate the interpolated temperatures and then replace the data in usable_hours with the new temperatures
                            distance_between_interpolation_temps += 1 # Accounts for both set temperatures used in calculating
                            interpolated_temperatures = [(beginning_interpolation_temp + (ending_interpolation_temp - beginning_interpolation_temp) * x / distance_between_interpolation_temps) for x in range(distance_between_interpolation_temps)]
                            interpolated_temperatures.pop(0) # Remove the first index which is the beginning_interpolation_temp
                            bitiy = beginning_interpolation_temp_index[0]
                            bitim = beginning_interpolation_temp_index[1]
                            bitid = beginning_interpolation_temp_index[2]
                            bitih = beginning_interpolation_temp_index[3]
                            eitiy = ending_interpolation_temp_index[0]
                            eitim = ending_interpolation_temp_index[1]
                            eitid = ending_interpolation_temp_index[2]
                            eitih = ending_interpolation_temp_index[3]
                            # If there was no day change, just replace the Nones between temperatures with the interpolated temperatures
                            if changed_days == False:
                                if loop_year < calc_years:
                                    self.usable_hourly[bitiy][bitim][bitid][bitih + 1 : eitih] = interpolated_temperatures
                                else:
                                    self.usable_hourly[bitiy][bitim][bitid][bitih + 1 : eitih] = interpolated_temperatures
                                    self.usable_hourly[-1][-1][-1].pop()
                            # If there was a day change, replace the temperatures after sundown of the first day and then the temperatures before sunup of the next day
                            else:
                                # Find the number of indexes between the beginning_interpolation_temp and the end of the day
                                end_of_day_distance = hours_per_day - bitih - 1
                                # Replace the end of the day with the necessary temperatures
                                self.usable_hourly[bitiy][bitim][bitid][bitih + 1:] = interpolated_temperatures[:end_of_day_distance]
                                # Replace the start of the next day with the necessary temperatures
                                self.usable_hourly[eitiy][eitim][eitid][:eitih] = interpolated_temperatures[end_of_day_distance:]
        # Third, include the impact from cloud cover
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        # Find cloud cover for the current hour and its impact
                        cloud_cover = cloud_param_data[year][month][day][hour]
                        if cloud_cover == 'Clear':
                            cloud_temperature_impact = random.choice([5, 6, 7])
                        elif cloud_cover == 'Mostly Clear':
                            cloud_temperature_impact = random.choice([2, 3, 4])
                        elif cloud_cover == 'Partly Cloudy':
                            cloud_temperature_impact = random.choice([-1, 0, 1])
                        elif cloud_cover == 'Mostly Cloudy':
                            cloud_temperature_impact = random.choice([-4, -3, -2])
                        else:
                            cloud_temperature_impact = random.choice([-7, -6, -5])
                        # Determine if the sun is up or not and adjust its impact based on how high in the sky the sun is
                        sunrise_hour = int(math.modf(sunrise_param_data[month][day])[1])
                        sunset_hour = int(math.modf(sundown_param_data[month][day])[1])
                        if hour < sunrise_hour or hour > sunset_hour:
                            cloud_temperature_impact_modifier = 0
                        else:
                            hours_of_sun = sunset_hour - sunrise_hour
                            max_hours = math.floor(hours_of_sun / 2)
                            hours_from_rise = hour - sunrise_hour
                            hours_to_set = sunset_hour - hour
                            if hours_from_rise <= hours_to_set:
                                hours_calc = hours_from_rise
                            else:
                                hours_calc = hours_to_set
                            cloud_temperature_impact_modifier = hours_calc / max_hours
                        cloud_temperature_impact_modifier = cloud_temperature_impact_modifier / 1.25 + 0.2
                        cloud_temperature_impact *= cloud_temperature_impact_modifier
                        # Apply cloud_temperature_impact to the existing temperature for this hour
                        self.usable_hourly[year][month][day][hour] += cloud_temperature_impact
                        # Lastly, round every temperature to two decimal places
                        rounded_hourly_temperature = round(self.usable_hourly[year][month][day][hour], 2)
                        self.usable_hourly[year][month][day][hour] = rounded_hourly_temperature
        print('Finished generate_temperatures')
        
class CloudData(BaseData):
    def __init__(self, clear, mostly_clear, partly_cloudy, mostly_cloudy, overcast):
        super().__init__()
        self.clear = self.interpolate_data(clear)
        self.mostly_clear = self.interpolate_data(mostly_clear)
        self.partly_cloudy = self.interpolate_data(partly_cloudy)
        self.mostly_cloudy = self.interpolate_data(mostly_cloudy)
        self.overcast = self.interpolate_data(overcast)
        
    def generate_clouds(self):
        print('Starting generate_clouds')
        # Determine cloud cover per hour
        cloud_system_list = []

        total_clouds = []
        for year in range(calc_years):
            yearly_clouds = []
            for month in range(months_per_year):
                monthly_clouds = []
                for day in range(days_per_month):
                    daily_clouds = []
                    for hour in range(hours_per_day):
                        # A cloud system (including clear skies) should always be active, so a new system should generate everytime there isn't one
                        if len(cloud_system_list) < 1:
                            # Determine what type of system based on the stacking odds of each type of system
                            cloud_rand = random.random() * 100
                            if cloud_rand < self.clear[month][day]:
                                cloud_system_type = 'Clear'
                            elif cloud_rand < self.clear[month][day] + self.mostly_clear[month][day]:
                                cloud_system_type = 'Mostly Clear'
                            elif cloud_rand < self.clear[month][day] + self.mostly_clear[month][day] + self.partly_cloudy[month][day]:
                                cloud_system_type = 'Partly Cloudy'
                            elif cloud_rand < self.clear[month][day] + self.mostly_clear[month][day] + self.partly_cloudy[month][day] + self.mostly_cloudy[month][day]:
                                cloud_system_type = 'Mostly Cloudy'
                            else:
                                cloud_system_type = 'Overcast'
                            # Determine the length of the system with weighted odds
                            cloud_system_type_list = [x + 1 for x in range(12)]
                            cloud_system_length_odds = self.create_weighted_list_exponential(12, 5, 1.2, 1.2)
                            cloud_system_length = random.choices(cloud_system_type_list, weights = cloud_system_length_odds, k = 1)[0]
                            # Populate the current cloud system list with the cloud system type based on length of the cloud system
                            for x in range(cloud_system_length):
                                cloud_system_list.append(cloud_system_type)
                            # Append the popped zero index of the cloud_system_list to the daily_clouds list
                            daily_clouds.append(cloud_system_list.pop(0))
                        else:
                            daily_clouds.append(cloud_system_list.pop(0))
                    monthly_clouds.append(daily_clouds)
                yearly_clouds.append(monthly_clouds)
            total_clouds.append(yearly_clouds)
        self.usable_hourly = total_clouds
        print('Finished generate_clouds')
                            
class PrecipitationData(BaseData):
    def __init__(self, precipitation):
        super().__init__()
        self.precipitation = self.interpolate_data(precipitation)

    def generate_precipitation(self):
        print('Starting generate_precipitation')
        # Determine which hours should have precipitation - Target precipitation frequency should be 0.33
        active_storm = False

        list_storm_strength = list(range(1,11))
        odds_storm_strength = [40, 60, 80, 100, 80, 60, 40, 30, 20, 10]

        list_storm_length = list(range(1, 25))
        odds_storm_length_1 = [100, 150, 200, 153.85, 102.56, 68.38, 45.58, 30.39, 20.26, 13.51, 9, 6, 4, 2.67, 1.78, 1.19, 0.79, 0.53, 0.35, 0.23, 0.16, 0.1, 0.07, 0.05]
        odds_storm_length_2_to_4 = [50, 100, 150, 200, 166.67, 138.89, 115.74, 96.45, 80.38, 66.98, 55.82, 46.51, 38.76, 32.3, 26.92, 22.43, 18.69, 15.58, 12.98, 10.82, 9.01, 7.51, 6.26, 5.22]
        odds_storm_length_5_to_7 = [20, 60, 100, 140, 180, 156.52, 136.11, 118.35, 102.92, 89.49, 77.82, 67.67, 58.84, 51.17, 44.49, 38.69, 33.64, 29.26, 25.44, 22.12, 19.24, 16.73, 14.54, 12.65]
        odds_storm_length_8_to_10 = [10, 70, 130, 190, 250, 310, 281.82, 256.2, 232.91, 211.73, 192.49, 174.99, 159.08, 144.62, 131.47, 119.52, 108.65, 98.78, 89.8, 81.63, 74.21, 67.47, 61.33, 55.76]

        total_precipitation_strength = []
        for year in range(calc_years):
            yearly_precipitation_strength = []
            for month in range(months_per_year):
                monthly_precipitation_strength = []
                for day in range(days_per_month):
                    daily_precipitation_strength = []
                    for hour in range(hours_per_day):
                        if active_storm == False:
                            # Populate total_precipitation with a zero to note that there is no current storm
                            daily_precipitation_strength.append(0)
                            # Determine if a storm should occur at the current hour - Divided by 4 to get the correct rate of occurrence
                            storm_rand = random.random() * 100
                            if storm_rand < self.precipitation[month][day] / 4:
                                active_storm = True
                                # Determine strength of the storm
                                storm_strength = random.choices(list_storm_strength, weights = odds_storm_strength, k = 1)[0]
                                # Determine length of storm based on the strength of the storm
                                if storm_strength <= 1:
                                    temp_odds_length = odds_storm_length_1.copy()
                                elif storm_strength <= 4:
                                    temp_odds_length = odds_storm_length_2_to_4.copy()
                                elif storm_strength <= 7:
                                    temp_odds_length = odds_storm_length_5_to_7.copy()
                                else:
                                    temp_odds_length = odds_storm_length_8_to_10.copy()
                                storm_length = random.choices(list_storm_length, weights = temp_odds_length, k = 1)[0]
                                temp_length_list = [x + 1 for x in range(storm_length)]
                                # Determine when the peak strength of the storm should occur
                                peak_odds_index = math.floor(storm_length / 4)
                                temp_odds_peak = self.create_weighted_list_exponential(storm_length, peak_odds_index, 1.2, 1.3)
                                storm_peak = random.choices(temp_length_list, weights = temp_odds_peak, k = 1)[0] - 1
                                # Determine the strength of each hour of the storm
                                storm_list = [0 for x in range(storm_length)]
                                storm_list[storm_peak] = storm_strength
                                storm_list_front = storm_list[:storm_peak + 1]
                                storm_list_back = storm_list[storm_peak:]
                                divisor_list = list(range(10, 51))
                                divisor_list = [x / 10 for x in divisor_list]
                                divisor_list_odds = self.create_weighted_list_exponential(len(divisor_list), 5, 1.2, 1.2)
                                for x in range(len(storm_list_front) - 1):
                                    divisor = random.choices(divisor_list, weights = divisor_list_odds, k = 1)[0]
                                    storm_list_front[-(x+2)] = storm_list_front[-(x+1)] / divisor
                                for x in range(len(storm_list_back) - 1):
                                    divisor = random.choices(divisor_list, weights = divisor_list_odds, k = 1)[0]
                                    storm_list_back[x+1] = storm_list_back[x] / divisor
                                storm_list_front.pop()
                                storm_list = storm_list_front + storm_list_back
                                storm_list = [math.ceil(x) for x in storm_list]
                        else:
                            if len(storm_list) > 1:
                                daily_precipitation_strength.append(storm_list.pop(0))
                            else:
                                active_storm = False
                                daily_precipitation_strength.append(0)
                    monthly_precipitation_strength.append(daily_precipitation_strength)
                yearly_precipitation_strength.append(monthly_precipitation_strength)
            total_precipitation_strength.append(yearly_precipitation_strength)
        self.usable_hourly_strength = total_precipitation_strength
        print('Finished generate_precipitation')

    def generate_precipitation_type(self, temperature_param_data, wind_speed_param_data):
        print('Starting generate_precipitation_type')
        # Determine what type of weather event each precipitation should be
        # Based on temperature, wind speed, and precipitation strength - Shifting scale of event likelihood based on temperature and strength
        # A temperature below 16 is always snow and above 32 is always rain - in between there is a chance of sleet or freezing rain
        # Wind used to determine snowstorm type - Snow or blizzard
        thunderstorm_likelihood_modifier_by_month = [0.1, 0.3, 0.5, 0.7, 0.9, 1, 1, 0.9, 0.7, 0.5, 0.3, 0.1]
        total_precipitation_type = []
        for year in range(calc_years):
            yearly_precipitation_type = []
            for month in range(months_per_year):
                monthly_precipitation_type = []
                for day in range(days_per_month):
                    daily_precipitation_type = []
                    for hour in range(hours_per_day):
                        current_storm_strength = self.usable_hourly_strength[year][month][day][hour]
                        # Determine if a precipitation event is active or not
                        if current_storm_strength > 0:
                            current_hour_temperature = temperature_param_data[year][month][day][hour]
                            # First, determine broad type of weather event based on temperature - Snow / Sleet / Freezing Rain / Rain
                            if current_hour_temperature < 16:
                                weather_event_broad_type = 'Snow'
                            elif current_hour_temperature >= 16 and current_hour_temperature <= 32:
                                # If the temperature is between 16 and 32, determine if it is Snow or an unusual weather event (Sleet / Freezing Rain)
                                snow_rand = random.random() * 100
                                if snow_rand < 10:
                                    unusual_event_rand = random.random() * 16
                                    if unusual_event_rand <= current_hour_temperature - 16:
                                        weather_event_broad_type = 'Sleet'
                                    else:
                                        weather_event_broad_type = 'Freezing Rain'
                                else:
                                    weather_event_broad_type = 'Snow'
                            else:
                                weather_event_broad_type = 'Rain'
                            # Second, if it is Snow or Rain determine if it should be changed to a Blizzard or a Thunderstorm
                            if weather_event_broad_type == 'Snow' and current_storm_strength > 4 and wind_speed_param_data[year][month][day][hour] >= 35:
                                weather_event = 'Blizzard'
                            elif weather_event_broad_type == 'Rain':
                                thunder_rand = random.random() * 11 # 11 is used for there to still be a slight chance of a 10 strength storm with no thunder
                                if thunder_rand <= current_storm_strength:
                                    if random.random() <= thunderstorm_likelihood_modifier_by_month[month]:
                                        weather_event = 'Thunderstorm'
                                    else:
                                      weather_event = 'Rain'
                                else:
                                    weather_event = 'Rain'
                            else:
                                weather_event = weather_event_broad_type
                        else:
                            weather_event = 0
                        daily_precipitation_type.append(weather_event)
                    monthly_precipitation_type.append(daily_precipitation_type)
                yearly_precipitation_type.append(monthly_precipitation_type)
            total_precipitation_type.append(yearly_precipitation_type)
        self.usable_hourly_type = total_precipitation_type
        print('Finished generate_precipitation_type')

class DewData(BaseData):
    def __init__(self, dew_point):
        super().__init__()
        self.dew_point = self.interpolate_data(dew_point)

    def generate_dew_points(self, temperature_param_data, temperature_day_average, precipitation_param_data):
        print('Starting generate_dew_points')
        # Dew point is generated based on: The average daily dew point in self.dew, existing generated precipitation, and a wandering variation
        # The wandering variation should be designed to make the hourly dew point closer to max before and after a storm and then more random between storms
        wandering_variation = 0
        wandering_correctiveness = 1 # Must be positive. A larger number means the temperature is more quickly corrected towards the average
        wandering_step_likelihood = 1.1 # Must be positive. A larger number means the temperature is less likely to wander by large amounts at once
        wandering_adjusted_peak = 12 # Must be between 0 and 24. Closer to 0 leads to an increased likelihood of the wandering variation to be decreased
        wandering_result_list = [x for x in range(-12, 13)]
        wandering_result_list = [x/4 for x in wandering_result_list]
        hours_since_storm = 0
        total_dew_point = []
        for year in range(calc_years):
            yearly_dew_point = []
            for month in range(months_per_year):
                monthly_dew_point = []
                for day in range(days_per_month):
                    daily_dew_point = []
                    # Recalculate wandering_result_list every day so the rate of the dew point change can be adjusted by the day's average temperature
                    wandering_result_list = [x for x in range(-24, 25)]
                    wandering_result_list_modifier = (temperature_day_average[month][day] / 100)
                    wandering_result_list = [x * wandering_result_list_modifier for x in wandering_result_list]
                    for hour in range(hours_per_day):
                        # If there is an active storm, set the wandering variation to the difference between the day's average dew point and the current temperature
                        if precipitation_param_data[year][month][day][hour] > 0:
                            wandering_variation = temperature_param_data[year][month][day][hour] - self.dew_point[month][day]
                            hours_since_storm = 0
                        # If there is no active storm, adjust the wandering variation based on how near the next storm is
                        else:
                            # Calculate number of hours until the next storm and the time since the last storm
                            hours_until_storm = 0
                            if hours_since_storm < 12:
                                hours_since_storm += 1
                            loop_year = year
                            loop_month = month
                            loop_day = day
                            loop_hour = hour
                            loops = 0
                            while True:
                                # Checks to see if 12 loops have already been gone through - since it doesn't matter if a storm is more than 12 hours away
                                if loops < 12:
                                    # Check if loop_year is at the end of the calc_years. If so, give a default value for hours_until_storm
                                    if loop_year < calc_years:
                                        # Check if loop_month needs to roll over to the next year
                                        if loop_month < months_per_year:
                                            # Check if loop_day needs to roll over to the next month
                                            if loop_day < days_per_month:
                                                # Check if loop_hour needs to roll over to the next day
                                                if loop_hour < hours_per_day:
                                                    if precipitation_param_data[loop_year][loop_month][loop_day][loop_hour] == 0:
                                                        hours_until_storm += 1
                                                        loop_hour += 1
                                                        loops += 1
                                                    else:
                                                        break
                                                else:
                                                    loop_hour = 0
                                                    loop_day += 1
                                            else:
                                                loop_day = 0
                                                loop_month += 1
                                        else:
                                            loop_month = 0
                                            loop_year += 1
                                    else:
                                        hours_until_storm = 12
                                        break
                                else:
                                    break
                            # Adjust the peak result of wondering_odds based on the hours until a storm occurs
                            adjusted_hours_until_storm = hours_until_storm - 24
                            adjusted_hours_since_storm = hours_since_storm - 24
                            wandering_adjusted_peak = 12 - adjusted_hours_until_storm + adjusted_hours_since_storm
                            adjusted_wandering_variation = (wandering_variation / 100) * wandering_correctiveness
                            if wandering_variation >= 0:
                                wandering_odds = self.create_weighted_list_exponential(49, wandering_adjusted_peak, wandering_step_likelihood - adjusted_wandering_variation, wandering_step_likelihood)
                            else:
                                wandering_odds = self.create_weighted_list_exponential(49, wandering_adjusted_peak, wandering_step_likelihood, wandering_step_likelihood + adjusted_wandering_variation)
                            wandering_variation += random.choices(wandering_result_list, weights = wandering_odds, k = 1)[0]
                            # Reduce wandering_variation to slightly less than 100% relative humidity if a storm is not active
                            if wandering_variation + self.dew_point[month][day] > temperature_param_data[year][month][day][hour] * 0.98:
                                wandering_variation = temperature_param_data[year][month][day][hour] * 0.98 - self.dew_point[month][day]
                                wandering_variation = math.floor(wandering_variation)
                        # Round the new hourly dew point and append it to the list
                        rounded_hourly_dew = round(self.dew_point[month][day] + wandering_variation, 2)
                        daily_dew_point.append(rounded_hourly_dew)
                    monthly_dew_point.append(daily_dew_point)
                yearly_dew_point.append(monthly_dew_point)
            total_dew_point.append(yearly_dew_point)
        self.usable_hourly_dew_point = total_dew_point
        print('Finished generate_dew_points')

    def generate_dew_frost(self, temperature_param_data, precipitation_param_data, sunrise_param_data, sundown_param_data):
        print('Starting generate_dew_frost')
        # Determine when dew and frost should form based on temperature, dew point, and sunrise/sunset
        # Temperature is used to determine if the air is saturated enough to form dew or frost by comparing the dew point and temperature
        # Additionally, temperature is used to determined if it should be dew or frost that is formed
        # Sunrise and sunset are used to determine if common dew and frost should form. It should be rare for dew and frost to form outside of the night/morning hours
        dew_rare_event = 0
        last_dew_event = ''
        total_dew_frost = []
        for year in range(calc_years):
            yearly_dew_frost = []
            for month in range(months_per_year):
                monthly_dew_frost = []
                for day in range(days_per_month):
                    daily_dew_frost = []
                    for hour in range(hours_per_day):
                        current_temperature = temperature_param_data[year][month][day][hour]
                        storm_active = True if precipitation_param_data[year][month][day][hour] > 0 else False
                        sunrise_hour = int(math.modf(sunrise_param_data[month][day])[1])
                        sunset_hour = int(math.modf(sundown_param_data[month][day])[1])
                        current_dew_point = self.usable_hourly_dew_point[year][month][day][hour]
                        # First, check if there is an active rare dew event and the current temperature is below 32 (A frost event)
                        if dew_rare_event > 0 and current_temperature < 32:
                            dew_frost_event = True
                            dew_rare_event -= 1
                        # Second, check if there is an active rare dew event and there is rain. If so, disable the active rare dew event and have no dew
                        elif dew_rare_event > 0 and storm_active:
                            dew_frost_event = False
                            dew_rare_event = 0
                        # Third, check if a precipitation event is active and there isn't a rare dew event active
                        elif storm_active:
                            dew_frost_event = False
                        # Fourth, check if the dew point and temperature are close together enough for dew and frost to form
                        elif current_dew_point >= current_temperature * 0.85:
                            # Check if it is in the night/morning and dew/frost should form else check if a rare daytime dew/frost should form
                            if hour < sunrise_hour + 2 or hour > sunset_hour + 3:
                                dew_frost_event = True
                            else:
                                dew_rare_event_rand = random.random() * 100
                                # In the event of a rare dew event, make it last at least 3 hours.
                                if dew_rare_event_rand <= 5:
                                    dew_rare_event = 3
                                    dew_frost_event = True
                        else:
                            dew_frost_event = False
                        # Next, determine if the dew frost event is dew or frost by looking at the current temperature
                        if dew_frost_event:
                            if current_temperature < 32 or (last_dew_event == 'Frost' and current_temperature < 36): # last_dew_event is used to prevent flip flopping
                                current_dew_frost_event, last_dew_event = 'Frost', 'Frost'
                            else:
                                current_dew_frost_event, last_dew_event = 'Dew', 'Dew'
                        else:
                            current_dew_frost_event, last_dew_event = '', ''
                        daily_dew_frost.append(current_dew_frost_event)
                    monthly_dew_frost.append(daily_dew_frost)
                yearly_dew_frost.append(monthly_dew_frost)
            total_dew_frost.append(yearly_dew_frost)
        self.usable_hourly_dew_event = total_dew_frost
        print('Finished generate_dew_frost')

    def generate_fog(self, temperature_param_data, precipitation_type_param_data):
        print('Starting generate_fog')
        # Determine if fog should be present based on dew point, temperature, and time until next rain storm / since last rain storm
        # Fog should form when dew point and temperature are close together
        # Fog should not form during rain storms but should form during some snow storms
        hours_since_storm = 0
        hours_until_storm = 0
        fog_during_precipitation = 0
        upcoming_storm = 'Default'
        total_fog = []
        for year in range(calc_years):
            yearly_fog = []
            for month in range(months_per_year):
                monthly_fog = []
                for day in range(days_per_month):
                    daily_fog = []
                    for hour in range(hours_per_day):
                        # If there is a fog_during_precipitation event, always have fog
                        if fog_during_precipitation > 0:
                            fog_during_precipitation -= 1
                            fog = 'Fog'
                        else:
                            # Check if the dew point is close enough to the temperature for fog to form
                            if self.usable_hourly_dew_point[year][month][day][hour] >= temperature_param_data[year][month][day][hour] * 0.95:
                                # If a storm is active, set hours_since_storm back to 0
                                if precipitation_type_param_data[year][month][day][hour] != 0:
                                    hours_since_storm = 0
                                # If a storm isn't active, add one to hours_since_storm and calculate how long until the next storm if it is unknown
                                else:
                                    hours_since_storm += 1
                                    if hours_until_storm > 0:
                                        hours_until_storm -= 1
                                    else:
                                        loop_year = year
                                        loop_month = month
                                        loop_day = day
                                        loop_hour = hour
                                        while True:
                                            # Check if loop_year is at the end of the calc_years. If so, give a default value for ending_interpolation_temp
                                            if loop_year < calc_years:
                                                # Check if loop_month needs to roll over to the next year
                                                if loop_month < months_per_year:
                                                    # Check if loop_day needs to roll over to the next month
                                                    if loop_day < days_per_month:
                                                        # Check if loop_hour needs to roll over to the next day
                                                        if loop_hour < hours_per_day:
                                                            if precipitation_type_param_data[loop_year][loop_month][loop_day][loop_hour] == 0:
                                                                hours_until_storm += 1
                                                                loop_hour += 1
                                                            else:
                                                                upcoming_storm = precipitation_type_param_data[loop_year][loop_month][loop_day][loop_hour]
                                                                break
                                                        else:
                                                            loop_hour = 0
                                                            loop_day += 1
                                                    else:
                                                        loop_day = 0
                                                        loop_month += 1
                                                else:
                                                    loop_month = 0
                                                    loop_year += 1
                                            else:
                                                break
                                if hours_since_storm > 3 and hours_until_storm > 3:
                                    fog = 'Fog'
                                else:
                                    fog_rand = random.random() * 100
                                    if fog_rand < 50 and upcoming_storm != 'Rain' and upcoming_storm != 'Thunderstorm':
                                        fog = 'Fog'
                                        fog_during_precipitation = random.choice([2, 3, 4])
                                    elif fog_rand < 15:
                                        fog = 'Fog'
                                        fog_during_precipitation = random.choice([2, 3, 4])
                                    else:
                                        fog = ''
                            else:
                                fog = ''
                        daily_fog.append(fog)
                    monthly_fog.append(daily_fog)
                yearly_fog.append(monthly_fog)
            total_fog.append(yearly_fog)
        self.usable_hourly_fog = total_fog
        print('Finished generate_fog')

class SunData(BaseData):
    def __init__(self, sunrise, sunset):
        super().__init__()
        self.sunrise = self.interpolate_data(sunrise)
        self.sunset = self.interpolate_data(sunset)

    def generate_sunrise_sunset(self):
        print('Starting generate_sunrise_sunset')
        # Sunrises and sunsets will always follow the same pattern - Only need to determine one year
        # Presented as 0 for sun is down, 1 for sun is up, and a decimal for when in that hour the sun rose or set
        yearly_sun = []
        yearly_sun_readable = []
        for month in range(months_per_year):
            monthly_sun_readable = []
            monthly_sun = []
            for day in range(days_per_month):
                daily_sun_readable = []
                daily_sun = []
                for hour in range(hours_per_day):
                    # Check to see if the sunrise is occurring in the current hour, if so, append the decimal hour
                    if hour <= self.sunrise[month][day] and hour + 1 >= self.sunrise[month][day]:
                        decimal_of_hour = math.modf(self.sunrise[month][day])[0]
                        decimal_of_hour = round(decimal_of_hour, 2)
                        daily_sun.append(decimal_of_hour)
                    # Check to see if the sunset is occuring in the current hour, if so, append the decimal hour.
                    elif hour <= self.sunset[month][day] and hour + 1 >= self.sunset[month][day]:
                        decimal_of_hour = math.modf(self.sunset[month][day])[0]
                        decimal_of_hour = round(decimal_of_hour, 2)
                        daily_sun.append(decimal_of_hour)
                    # Check to see if it is before sunrise, append a 0 for the sun being down
                    elif hour <= self.sunrise[month][day]:
                        daily_sun.append(0)
                    # Check to see if the hour is between sunrise and sunset, if so, append a 1 for the sun being up
                    elif hour >= self.sunrise[month][day] and hour < self.sunset[month][day]:
                        daily_sun.append(1)
                    # Check to see if it after sundown, append a 0 for the sun being down
                    elif hour >= self.sunset[month][day]:
                        daily_sun.append(0)
                # Convert the daily_sun data in a more readable format
                daily_sun_readable = daily_sun.copy()
                loop_hour = 0
                hours_of_sunlight = 0
                total_degrees = 0 # Used to track the position of the sun in the sky at the start of each hour
                while loop_hour < hours_per_day:
                    if daily_sun_readable[loop_hour] == 1:
                        hours_of_sunlight += 1
                    elif daily_sun_readable[loop_hour] > 0:
                        hours_of_sunlight += 1 - daily_sun_readable[loop_hour]
                    loop_hour += 1
                degrees_per_hour = 180 / hours_of_sunlight # How far across the sky the sun travels in an hour
                for hour in range(hours_per_day):
                    if daily_sun_readable[hour] == 0:
                        daily_sun_readable[hour] = ''
                    else:
                        if daily_sun_readable[hour - 1] == '':
                            total_degrees += (1 - daily_sun_readable[hour]) * degrees_per_hour
                            daily_sun_readable[hour] = f'{hour}:{round(daily_sun_readable[hour] * 60)}'
                        elif daily_sun_readable[hour] == 1:
                            daily_sun_readable[hour] = f'{round(total_degrees)}d'
                            total_degrees += degrees_per_hour
                        else:
                            daily_sun_readable[hour] = f'{hour}:{round(daily_sun_readable[hour] * 60)}'
                monthly_sun_readable.append(daily_sun_readable)
                monthly_sun.append(daily_sun)
            yearly_sun_readable.append(monthly_sun_readable)
            yearly_sun.append(monthly_sun)
        self.readable_sun_hourly = yearly_sun_readable
        self.usable_sun_decimal_hourly = yearly_sun
        print('Finished generate_sunrise_sunset')
    
    def generate_moonrise_moonset(self):
        print('Starting generate_moonrise_moonset')
        # Moon should cycle should be a full 28 days - e.g. Full moon -> 14 days -> New Moon -> 14 Days -> Full Moon
        # Moonset on a full moon should be roughly equal to the moonrise on a new moon       
        unadjusted_moonrise = deepcopy(self.sunset)
        unadjusted_moonset = deepcopy(self.sunrise)
        moonrise = []
        moonset = []
        for month in range(months_per_year):
            monthly_moonrise = []
            monthly_moonset = []
            for day in range(days_per_month):
                monthly_moonrise.append(unadjusted_moonrise[month][day] - 1)
                monthly_moonset.append(unadjusted_moonset[month][day] + 1)
            moonrise.append(monthly_moonrise)
            moonset.append(monthly_moonset)
        yearly_moon = []
        for month in range(months_per_year):
            monthly_moon = []
            for day in range(days_per_month):
                daily_moon = []
                for hour in range(hours_per_day):
                    # Check to see if the moonrise is occurring in the current hour, if so, append the decimal hour
                    if hour <= moonrise[month][day] and hour + 1 >= moonrise[month][day]:
                        decimal_of_hour = math.modf(moonrise[month][day])[0]
                        decimal_of_hour = round(decimal_of_hour, 2)
                        daily_moon.append(decimal_of_hour)
                    # Check to see if the moonset is occuring in the current hour, if so, append the decimal hour.
                    elif hour <= moonset[month][day] and hour + 1 >= moonset[month][day]:
                        decimal_of_hour = math.modf(moonset[month][day])[0]
                        decimal_of_hour = round(decimal_of_hour, 2)
                        daily_moon.append(decimal_of_hour)
                    # Check to see if it is before moonset, append a 1 for the moon being up
                    elif hour <= moonset[month][day]:
                        daily_moon.append(1)
                    # Check to see if the hour is between moonset and moonrise, if so, append a 0 for the moon being down
                    elif hour >= moonset[month][day] and hour < moonrise[month][day]:
                        daily_moon.append(0)
                    # Check to see if it after moonrise, append a 1 for the moon being up
                    elif hour >= moonrise[month][day]:
                        daily_moon.append(1)
                monthly_moon.append(daily_moon)
            yearly_moon.append(monthly_moon)
        self.usable_moon_decimal_hourly = yearly_moon
        hours_of_moonlight = 0
        total_degrees = 0
        yearly_moon_readable = []
        for month in range(months_per_year):
            monthly_moon_readable = []
            for day in range(days_per_month):
                daily_moon_readable = []
                for hour in range(hours_per_day):
                    # Get current, next, and prior hour amounts
                    current_hour_amount = self.usable_moon_decimal_hourly[month][day][hour]
                    if hour == hours_per_day - 1:
                        next_hour_amount = 1
                    else:
                        next_hour_amount = self.usable_moon_decimal_hourly[month][day][hour + 1]
                    if hour == 0:
                        prior_hour_amount = 1
                    else:
                        prior_hour_amount = self.usable_moon_decimal_hourly[month][day][hour - 1]                           
                    # Feed default amounts for beginning and end indexes
                    if month == 0 and day == 0 and hour == 0:
                        hours_of_moonlight = 13.34
                        degrees_per_hour = 180 / hours_of_moonlight
                        total_degrees = 6.74 * degrees_per_hour
                        daily_moon_readable.append(f'{round(total_degrees)}d')
                    elif month == months_per_year - 1 and day == days_per_month - 1 and hour == hours_per_day - 1:
                        daily_moon_readable.append(f'{round(total_degrees)}d')
                    # Cycle through and determine hours_of_moonlight every time a non-0 is discovered after a 0 and create the moonrise time
                    elif current_hour_amount > 0 and prior_hour_amount == 0:
                        loop_month = month
                        loop_day = day
                        loop_hour = hour
                        while True:
                            # Check if loop_month is at the end of the months_per_year. If so, give a default value for hours_of_moonlight
                            if loop_month < months_per_year:
                                # Check if loop_day needs to roll over to the next month
                                if loop_day < days_per_month:
                                    # Check if loop_hour needs to roll over to the next day
                                    if loop_hour < hours_per_day:
                                        if self.usable_moon_decimal_hourly[loop_month][loop_day][loop_hour] == 1:
                                            hours_of_moonlight += 1
                                            loop_hour += 1
                                        elif self.usable_moon_decimal_hourly[loop_month][loop_day][loop_hour] > 0:
                                            hours_of_moonlight += 1 - self.usable_moon_decimal_hourly[loop_month][loop_day][loop_hour]
                                            loop_hour += 1
                                        else:
                                            if day == 27:
                                                moon_fullness = 100
                                                daily_moon_readable[hour - 1] = f'{moon_fullness}% - Full'
                                            elif day == 13:
                                                moon_fullness = 0
                                                daily_moon_readable[hour - 1] = f'{moon_fullness}% - New'
                                            elif day < 14:
                                                moon_fullness = round((1-((day+1)%14/14))*100)
                                                daily_moon_readable[hour - 1] = f'{moon_fullness}% - Waning'
                                            else:
                                                moon_fullness = round(((day+1)%14/14)*100)
                                                daily_moon_readable[hour - 1] = f'{moon_fullness}% - Waxing'
                                            daily_moon_readable.append(f'{hour}:{round(current_hour_amount * 60)}')
                                            degrees_per_hour = 180 / hours_of_moonlight
                                            total_degrees += degrees_per_hour * (1 - current_hour_amount)
                                            break
                                    else:
                                        loop_hour = 0
                                        loop_day += 1
                                else:
                                    loop_day = 0
                                    loop_month += 1
                            else:
                                hours_of_moonlight = 13.34
                                daily_moon_readable.append(f'{hour}:{round(current_hour_amount * 60)}')
                                degrees_per_hour = 180 / hours_of_moonlight
                                total_degrees += degrees_per_hour * (1 - current_hour_amount)
                                break
                    # If a 0 is found after a non-0 number, set hours_of_moonlight and total_degrees back to 0
                    elif current_hour_amount > 0 and next_hour_amount == 0:
                        daily_moon_readable.append(f'{hour}:{round(current_hour_amount * 60)}')
                        hours_of_moonlight = 0
                        total_degrees = 0
                    # If a 1 is found, show the sun as up and its location in the sky
                    elif current_hour_amount == 1:
                        daily_moon_readable.append(f'{round(total_degrees)}d')
                        total_degrees += degrees_per_hour
                    else:
                        daily_moon_readable.append('')
                monthly_moon_readable.append(daily_moon_readable)
            yearly_moon_readable.append(monthly_moon_readable)
        self.readable_moon_hourly = yearly_moon_readable
        print('Finished generate_moonrise_moonset')

class WindData(BaseData):
    def __init__(self, wind_speed, wind_dir_N, wind_dir_E, wind_dir_S, wind_dir_W):
        super().__init__()
        self.wind_speed = self.interpolate_data(wind_speed)
        self.wind_dir_N = self.interpolate_data(wind_dir_N)
        self.wind_dir_E = self.interpolate_data(wind_dir_E)
        self.wind_dir_S = self.interpolate_data(wind_dir_S)
        self.wind_dir_W = self.interpolate_data(wind_dir_W)
    
    def generate_wind_speeds(self, precipitation_param_data, sun_param_data):
        print('Starting generate_wind_speeds')
        # Wind speed is based on daily average wind speed, if the sun is up, and strength of precipitation
        # A greater variation in wind while the sun is up
        # An increased chance of strong windows during high precipitation
        wind_speed_wandering_variation = 0
        wind_speed_wandering_correctiveness = 1.5 # Must be positive. A larger number means the wind speed is more quickly corrected towards the average
        wind_speed_wandering_step_likelihood = 1.4 # Must be positive. A larger number means the wind speed is less likely to wander by large amounts at once
        wind_speed_wandering_result_list = [x for x in range(-12, 13)]
        wind_speed_wandering_result_list = [x/2 for x in wind_speed_wandering_result_list]
        storm_strength_step_likelihood = 1.4 # Must be positive. A larger number means the wind speed is less likely to differ by large amounts compared to the same strength storm
        storm_strength_result_list = [x for x in range(51)]
        storm_strength_adjusted_peak = 0 # Must be between 0 and 51. A higher peak leads to an increased chance of strong winds during a storm
        num_of_choices = len(storm_strength_result_list)

        total_wind_speeds = []
        for year in range(calc_years):
            yearly_wind_speeds  = []
            for month in range(months_per_year):
                monthly_wind_speeds  = []
                for day in range(days_per_month):
                    daily_wind_speeds  = []
                    for hour in range(hours_per_day):
                        # If the sun is up, the wandering_variation should be more likely to wander
                        sun_up = sun_param_data[month][day][hour] != 0
                        wind_speed_wandering_step_likelihood = 1.2 if sun_up else 1.4
                        adjusted_wandering_variation = (wind_speed_wandering_variation / 100) * wind_speed_wandering_correctiveness
                        if wind_speed_wandering_variation >= 0:
                            wind_speed_wandering_odds = self.create_weighted_list_exponential(25, 12, wind_speed_wandering_step_likelihood - adjusted_wandering_variation, wind_speed_wandering_step_likelihood)
                        else:
                            wind_speed_wandering_odds = self.create_weighted_list_exponential(25, 12, wind_speed_wandering_step_likelihood, wind_speed_wandering_step_likelihood + adjusted_wandering_variation)
                        wind_speed_wandering_variation += random.choices(wind_speed_wandering_result_list, weights = wind_speed_wandering_odds, k = 1)[0]
                        # Set current wind speed to the average for the day + the wandering variation
                        # If the calculated temperature is below zero, set it to zero and help correct wandering_variation
                        current_wind_speed = self.wind_speed[month][day] + wind_speed_wandering_variation
                        if current_wind_speed < 0:
                            current_wind_speed = 0
                            wind_speed_wandering_variation = -self.wind_speed[month][day]
                        # Increase the wind speed if a storm is present and scale the likelihood of strong winds based on the strength of the storm
                        if precipitation_param_data[year][month][day][hour] > 0:
                            storm_strength_adjusted_peak = precipitation_param_data[year][month][day][hour] * 5
                            storm_strength_odds = self.create_weighted_list_exponential(num_of_choices, storm_strength_adjusted_peak, storm_strength_step_likelihood)
                            storm_strength_modifier = random.choices(storm_strength_result_list, weights = storm_strength_odds, k = 1)[0]
                            current_wind_speed += storm_strength_modifier
                        rounded_current_wind_speed = round(current_wind_speed, 2)
                        daily_wind_speeds.append(rounded_current_wind_speed)
                    monthly_wind_speeds.append(daily_wind_speeds)
                yearly_wind_speeds.append(monthly_wind_speeds)
            total_wind_speeds.append(yearly_wind_speeds)
        self.usable_hourly_speed = total_wind_speeds
        print('Finished generate_wind_speeds')

    def generate_wind_directions(self, dew_point_avg_param_data, dew_point_usable_param_data):
        print('Starting generate_wind_directions')
        # Wind direction is determined randomly but influenced by dew point
        wind_humdity_result_list = [x for x in range(0, 1080)] # Using 1080 to prevent issues with looping around the full 360 degrees
        wind_humidity_step_likelihood = 1.05
        wind_month_result_list = [x for x in range(-90, 91)]
        wind_month_step_likelihood = 1.05
        wind_system_length_result_list = [x for x in range (1,13)]
        wind_system_length_step_likelihood = 1.5
        wind_system_length = 0
        wind_system_direction = 0
        total_wind_directions = []
        for year in range(calc_years):
            yearly_wind_directions  = []
            for month in range(months_per_year):
                monthly_wind_directions  = []
                for day in range(days_per_month):
                    daily_wind_directions  = []
                    for hour in range(hours_per_day):
                        # First, check if there is an active wind system
                        if wind_system_length > 0:
                            wind_direction = wind_system_direction
                            wind_system_length -= 1
                        # If there isn't, create a new wind system
                        else:
                            # Determine the seasonal influence - More north in the fall/winter and more south in the spring/summer
                            if month >= 4 and month <= 9:
                                wind_direction_month_odds = self.create_weighted_list_exponential(len(wind_month_result_list), 45, wind_month_step_likelihood)
                            else:
                                wind_direction_month_odds = self.create_weighted_list_exponential(len(wind_month_result_list), 135, wind_month_step_likelihood)
                            wind_direction_month_influence = random.choices(wind_month_result_list, weights = wind_direction_month_odds, k = 1)[0]
                            # Determine if the air is more humid or less humid
                            humidity_difference = dew_point_usable_param_data[year][month][day][hour] - dew_point_avg_param_data[month][day]
                            # If the humidity is below average, dry air from the mountains comes in. If the humidity is above average, wet air from the ocean comes in
                            if humidity_difference >= 0:
                                wind_direction_humidity_odds = self.create_weighted_list_exponential(len(wind_humdity_result_list), 450 - wind_direction_month_influence, wind_humidity_step_likelihood)
                            else:
                                wind_direction_humidity_odds = self.create_weighted_list_exponential(len(wind_humdity_result_list), 630 + wind_direction_month_influence, wind_humidity_step_likelihood)
                            wind_direction_angle = random.choices(wind_humdity_result_list, weights = wind_direction_humidity_odds, k = 1)[0]
                            wind_system_length_odds = self.create_weighted_list_exponential(len(wind_system_length_result_list), 5, wind_system_length_step_likelihood)
                            wind_system_length = random.choices(wind_system_length_result_list, weights = wind_system_length_odds, k = 1)[0]
                            wind_direction_angle = wind_direction_angle % 360
                            if wind_direction_angle < 23:
                                wind_system_direction = 'S'
                            elif wind_direction_angle < 68:
                                wind_system_direction = 'SW'
                            elif wind_direction_angle < 113:
                                wind_system_direction = 'W'
                            elif wind_direction_angle < 158:
                                wind_system_direction = 'NW'
                            elif wind_direction_angle < 203:
                                wind_system_direction = 'N'
                            elif wind_direction_angle < 248:
                                wind_system_direction = 'NE'
                            elif wind_direction_angle < 293:
                                wind_system_direction = 'E'
                            elif wind_direction_angle < 338:
                                wind_system_direction = 'SE'
                            else:
                                wind_system_direction = 'S'
                            wind_direction = wind_system_direction
                            wind_system_length -= 1
                        daily_wind_directions.append(wind_direction)
                    monthly_wind_directions.append(daily_wind_directions)
                yearly_wind_directions.append(monthly_wind_directions)
            total_wind_directions.append(yearly_wind_directions)
        self.usable_hourly_direction = total_wind_directions
        print('Finished generate_wind_directions')

class BaseExcel:
    def __init__(self, workbook_file, worksheet_name):
        self.workbook_file = workbook_file
        self.workbook = load_workbook(workbook_file)
        self.worksheet = self.workbook[worksheet_name]
        self.date_column = 1
        self.start_row = 2
        self.rows_of_info_per_day = 9

    def populate_defaults(self, start_year):
        print('Starting populate_defaults')
        weather_event_labels = ['Temperature', 'Humidity', 'Storm Strength', 'Storm Type', 'Wind Speed', 'Wind Direction', 'Clouds', 'Sun', 'Moon']
        active_row = self.start_row
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    #### Date
                    # Merge and populate date name
                    self.worksheet.merge_cells(start_column = 1, start_row = active_row, end_column = 1, end_row = active_row + self.rows_of_info_per_day - 1)
                    merged_date_cell = self.worksheet.cell(column = 1, row = active_row)
                    merged_date_cell.value = f'{month + 1}/{day + 1}/{year + start_year}'
                    # Align merged date to center
                    merged_date_cell.alignment = Alignment(horizontal='center', vertical='center')
                    active_row += self.rows_of_info_per_day
                    #### Weather event labels
                    for label_index in range(len(weather_event_labels)):
                        self.worksheet.cell(column=self.date_column + 1, row=active_row + label_index).value = weather_event_labels[label_index]
                        self.worksheet.cell(column=self.date_column + 1, row=active_row + label_index).alignment = Alignment(horizontal='right')
        print('Finished populate_defaults')

    def populate_standard_row(self, row_to_populate, param_data):
        print('Starting populate_standard_row - ' + str(row_to_populate))
        active_row = self.start_row + row_to_populate - 1
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).value = param_data[year][month][day][hour]
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).style = 'Comma'
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).font = Font(name='Calibri', size=11)
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).alignment = Alignment(horizontal='center')
                    active_row += self.rows_of_info_per_day
        print('Finished populate_standard_row - ' + str(row_to_populate))

    def populate_precipitation_type_row(self, row_to_populate, precipitation_param_data):
        print('Starting populate_precipitation_type_row')
        active_row = self.start_row + row_to_populate - 1
        converted_names = deepcopy(precipitation_param_data)
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        if precipitation_param_data[year][month][day][hour] == 'Thunderstorm':
                            converted_names[year][month][day][hour] = 'Tstorm'
                        elif precipitation_param_data[year][month][day][hour] == 'Freezing Rain':
                            converted_names[year][month][day][hour] = 'Fr Rain'
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).value = converted_names[year][month][day][hour]
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).style = 'Comma'
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).font = Font(name='Calibri', size=11)
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).alignment = Alignment(horizontal='center')
                    active_row += self.rows_of_info_per_day
        print('Finished populate_precipitation_type_row')

    def populate_cloud_row(self, row_to_populate, cloud_param_data):
        print('Starting populate_cloud_row')
        active_row = self.start_row + row_to_populate - 1
        converted_names = deepcopy(cloud_param_data)
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        if cloud_param_data[year][month][day][hour] == 'Mostly Clear':
                            converted_names[year][month][day][hour] = 'M Clear'
                        elif cloud_param_data[year][month][day][hour] == 'Partly Cloudy':
                            converted_names[year][month][day][hour] = 'P Cloudy'
                        elif cloud_param_data[year][month][day][hour] == 'Mostly Cloudy':
                            converted_names[year][month][day][hour] = 'M Cloudy'
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).value = converted_names[year][month][day][hour]
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).style = 'Comma'
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).font = Font(name='Calibri', size=11)
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).alignment = Alignment(horizontal='center')
                    active_row += self.rows_of_info_per_day
        print('Finished populate_cloud_row')

    def populate_sun_moon_row(self, row_to_populate, sun_moon_param_data):
        print('Starting populate_sun_moon_row - ' + str(row_to_populate))
        active_row = self.start_row + row_to_populate - 1
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        if '-' in sun_moon_param_data[month][day][hour]:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).value = sun_moon_param_data[month][day][hour]
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).style = 'Comma'
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).font = Font(name='Calibri', size=11)
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).alignment = Alignment(horizontal='right')
                        else:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).value = sun_moon_param_data[month][day][hour]
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).style = 'Comma'
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).font = Font(name='Calibri', size=11)
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).alignment = Alignment(horizontal='center')
                    active_row += self.rows_of_info_per_day
        print('Finished populate_sun_moon_row - ' + str(row_to_populate))

    def color_scale(self, row_to_color_scale, param_data):
        print('Starting color_scale - ' + str(row_to_color_scale))
        num_min = param_data[0][0][0][0]
        num_max = param_data[0][0][0][0]
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        hour_temp = param_data[year][month][day][hour]
                        if hour_temp > num_max:
                            num_max = hour_temp
                        elif hour_temp < num_min:
                            num_min = hour_temp

        active_row = self.start_row + row_to_color_scale - 1
        min_color = FormatObject(type='num', val=num_min)
        max_color = FormatObject(type='num', val=num_max)
        if row_to_color_scale == 1:
            colors = [Color('00B0F0'), Color('FF6600')]
        elif row_to_color_scale == 2:
            colors = [Color('833C0C'), Color('9BC2E6')]
        elif row_to_color_scale == 5:
            colors = [Color('FFEF9C'), Color('FF7128')]
        color_scale = ColorScale(cfvo=[min_color, max_color], color=colors)
        temperature_rule = Rule(type='colorScale', colorScale=color_scale)
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    converted_selection = 'C' + str(active_row) + ':Z' + str(active_row)
                    self.worksheet.conditional_formatting.add(converted_selection, temperature_rule)
                    active_row += self.rows_of_info_per_day
        print('Finished color_scale - ' + str(row_to_color_scale))
        
    def precipitation_strength_fill(self, row_to_fill, precipitation_param_data):
        print('Starting precipitation_strength_fill')
        active_row = self.start_row + row_to_fill - 1
        whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        oneFill = PatternFill(start_color='a6b8da', end_color='a6b8da', fill_type='solid')
        twoFill = PatternFill(start_color='97aacd', end_color='97aacd', fill_type='solid')
        threeFill = PatternFill(start_color='889cc0', end_color='889cc0', fill_type='solid')
        fourFill = PatternFill(start_color='798db3', end_color='798db3', fill_type='solid')
        fiveFill = PatternFill(start_color='6a7fa6', end_color='6a7fa6', fill_type='solid')
        sixFill = PatternFill(start_color='5c7199', end_color='5c7199', fill_type='solid')
        sevenFill = PatternFill(start_color='4d628c', end_color='4d628c', fill_type='solid')
        eightFill = PatternFill(start_color='3e547f', end_color='3e547f', fill_type='solid')
        nineFill = PatternFill(start_color='2f4672', end_color='2f4672', fill_type='solid')
        tenFill = PatternFill(start_color='203764', end_color='203764', fill_type='solid')
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        current_strength = precipitation_param_data[year][month][day][hour]
                        match current_strength:
                            case 0:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = whiteFill
                            case 1:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = oneFill
                            case 2:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = twoFill
                            case 3:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = threeFill
                            case 4:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = fourFill
                            case 5:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = fiveFill
                            case 6:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = sixFill
                            case 7:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = sevenFill
                            case 8:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = eightFill
                            case 9:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = nineFill
                            case 10:
                                self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = tenFill
                    active_row += self.rows_of_info_per_day
        print('Finished precipitation_strength_fill')

    def precipitation_type_fill(self, row_to_fill, precipitation_param_data):
        print('Starting precipitation_type_fill')
        active_row = self.start_row + row_to_fill - 1
        whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        snowFill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        sleetFill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        frRainFill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
        rainFill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        tstormFill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        if precipitation_param_data[year][month][day][hour] == 0:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = whiteFill
                        elif precipitation_param_data[year][month][day][hour] == 'Snow':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = snowFill
                        elif precipitation_param_data[year][month][day][hour] == 'Sleet':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = sleetFill
                        elif precipitation_param_data[year][month][day][hour] == 'Freezing Rain':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = frRainFill
                        elif precipitation_param_data[year][month][day][hour] == 'Rain':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = rainFill
                        elif precipitation_param_data[year][month][day][hour] == 'Thunderstorm':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = tstormFill
                    active_row += self.rows_of_info_per_day
        print('Finished precipitation_type_fill')

    def wind_direction_fill(self, row_to_fill):
        print('Starting wind_direction_fill')
        active_row = self.start_row + row_to_fill - 1
        whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = whiteFill

    def cloud_fill(self, row_to_fill, cloud_param_data):
        print('Starting cloud_fill')
        active_row = self.start_row + row_to_fill - 1
        clearFill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        mClearFill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        pCloudyFill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
        mCloudyFill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
        overcastFill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        if cloud_param_data[year][month][day][hour] == 'Clear':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = clearFill
                        elif cloud_param_data[year][month][day][hour] == 'Mostly Clear':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = mClearFill
                        elif cloud_param_data[year][month][day][hour] == 'Partly Cloudy':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = pCloudyFill
                        elif cloud_param_data[year][month][day][hour] == 'Mostly Cloudy':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = mCloudyFill
                        elif cloud_param_data[year][month][day][hour] == 'Overcast':
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = overcastFill
                    active_row += self.rows_of_info_per_day
        print('Finished cloud_fill')

    def sun_fill(self, row_to_fill, sun_param_data):
        print('Starting sun_fill')
        active_row = self.start_row + row_to_fill - 1
        riseSetFill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        upFill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        downFill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        if ':' in sun_param_data[month][day][hour]:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = riseSetFill
                        elif 'd' in sun_param_data[month][day][hour]:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = upFill
                        else:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = downFill
                    active_row += self.rows_of_info_per_day
        print('Finished sun_fill')

    def moon_fill(self, row_to_fill, moon_param_data):
        print('Starting moon_fill')
        active_row = self.start_row + row_to_fill - 1
        riseSetFill = PatternFill(start_color='AEAAAA', end_color='AEAAAA', fill_type='solid')
        upFill = PatternFill(start_color='757171', end_color='757171', fill_type='solid')
        downFill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    for hour in range(hours_per_day):
                        if ':' in moon_param_data[month][day][hour]:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = riseSetFill
                        elif 'd' in moon_param_data[month][day][hour]:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = upFill
                        else:
                            self.worksheet.cell(column=self.date_column + 2 + hour, row=active_row).fill = downFill
                    active_row += self.rows_of_info_per_day
        print('Finished moon_fill')

    def populate_borders(self):
        print('Starting populate_borders')
        medium = Side(border_style='medium', color='000000')
        active_row = self.start_row
        for year in range(calc_years):
            for month in range(months_per_year):
                for day in range(days_per_month):
                    # Separate dates with a border
                    column = 1
                    if active_row != self.start_row:
                        while column <= 26:
                            self.worksheet.cell(column=column, row=active_row).border = Border(top=medium)
                            column += 1
                    active_row += self.rows_of_info_per_day
        print('Finished populate_borders')

    def save_workbook(self):
        print('Saving Workbook')
        self.workbook.save(self.workbook_file)
        print('Workbook saved')

#### Global Variables
# Number of days in a month and years to calculate
calc_years = 10
months_per_year = 12
days_per_month = 28
hours_per_day = 24

##### Raw Data
# Data points given by beginning of each month with the first repeated at the end e.g. [Jan1, Feb1, Mar1, ..., Jan1]
# Temp given as fahrenheit
raw_temp_avg_low = [29, 30, 37, 46, 56, 65, 71, 69, 62, 52, 42, 34, 29]
raw_temp_avg_avg = [33, 35, 42, 53, 63, 72, 77, 76, 69, 58, 48, 39, 33]
raw_temp_avg_high = [40, 42, 51, 61, 71, 79, 84, 82, 75, 65, 54, 45, 40]

# Clouds given as percent chance of each type
raw_clouds_clear = [30, 28, 28, 30, 30, 28, 32, 35, 42, 43, 39, 31, 30]
raw_clouds_mostly_clear = [10, 10, 10, 10, 10, 10, 10, 8, 7, 8, 9, 10, 10]
raw_clouds_partly_cloudy = [8, 10, 10, 9, 10, 11, 12, 12, 10, 9, 9, 8, 8]
raw_clouds_mostly_cloudy = [10, 10, 10, 10, 10, 13, 13, 14, 10, 10, 10, 10, 10]
raw_clouds_overcast = [42, 42, 42, 41, 40, 38, 33, 31, 31, 30, 33, 41, 42]

# Preciptation given as daily percent chance of precipitation
raw_precipitation_avg = [26, 24, 30, 31, 35, 35, 37, 35, 28, 26, 27, 29, 26]

# Dew point given as dew point in fahrenheit
raw_dew_point_avg = [17.9, 18.3, 24.2, 34.7, 47.9, 57.3, 62.7, 61.3, 54.7, 44.8, 32.7, 24.9, 17.9]

# Sunrise and Sunset given as hour decimal format
raw_sunrise = [7.58, 7.32, 6.7, 5.82, 5.05, 4.57, 4.58, 5.0, 5.53, 6.07, 6.65, 7.25, 7.58]
raw_sunset = [16.77, 17.37, 17.95, 19.38, 19.1, 19.6, 19.77, 19.42, 18.67, 17.8, 17.02, 16.6, 16.77]

# Wind speed is given as an average in mph
raw_wind_speed = [10.1, 10.2, 9.8, 8.7, 7.4, 6.7, 6.4, 6.4, 7.3, 8.4, 9.2, 9.7, 10.1]

# Wind direction is given as percent chance of direction
raw_wind_direction_north = [28, 32, 32, 30, 25, 21, 19, 20, 25, 27, 30, 28, 28]
raw_wind_direction_east = [10, 10, 13, 18, 18, 18, 11, 12, 17, 17, 13, 12, 10]
raw_wind_direction_south = [16, 16, 18, 21, 24, 28, 33, 35, 36, 28, 22, 18, 16]
raw_wind_direction_west = [46, 42, 37, 31, 33, 33, 37, 33, 22, 28, 35, 42, 46]

##### Declare data classes and interpolate the data
temperature_data = TemperatureData(raw_temp_avg_low, raw_temp_avg_avg, raw_temp_avg_high)
cloud_data = CloudData(raw_clouds_clear, raw_clouds_mostly_clear, raw_clouds_partly_cloudy, raw_clouds_mostly_cloudy, raw_clouds_overcast)
precipitation_data = PrecipitationData(raw_precipitation_avg)
sun_moon_data = SunData(raw_sunrise, raw_sunset)
dew_data = DewData(raw_dew_point_avg)
wind_data = WindData(raw_wind_speed, raw_wind_direction_north, raw_wind_direction_east, raw_wind_direction_south, raw_wind_direction_west)

#### Order of data calculations
precipitation_data.generate_precipitation()
cloud_data.generate_clouds()
sun_moon_data.generate_sunrise_sunset()
sun_moon_data.generate_moonrise_moonset()
temperature_data.generate_temperatures(sun_moon_data.sunrise, sun_moon_data.sunset, cloud_data.usable_hourly)
dew_data.generate_dew_points(temperature_data.usable_hourly, temperature_data.avg, precipitation_data.usable_hourly_strength)
wind_data.generate_wind_speeds(precipitation_data.usable_hourly_strength, sun_moon_data.usable_sun_decimal_hourly)
precipitation_data.generate_precipitation_type(temperature_data.usable_hourly, wind_data.usable_hourly_speed)
dew_data.generate_dew_frost(temperature_data.usable_hourly, precipitation_data.usable_hourly_strength, sun_moon_data.sunrise, sun_moon_data.sunset)
dew_data.generate_fog(temperature_data.usable_hourly, precipitation_data.usable_hourly_type)
wind_data.generate_wind_directions(dew_data.dew_point, dew_data.usable_hourly_dew_point)

#### Load workbook
workbook = 'sim.xlsx'
worksheet = 'Data'

#### Declare Excel classes
base_excel = BaseExcel(workbook, worksheet)

#### Order of Excel population
base_excel.populate_defaults(3304)
base_excel.populate_standard_row(1, temperature_data.usable_hourly)
base_excel.populate_standard_row(2, dew_data.usable_hourly_dew_point)
base_excel.populate_standard_row(3, precipitation_data.usable_hourly_strength)
base_excel.populate_precipitation_type_row(4, precipitation_data.usable_hourly_type)
base_excel.populate_standard_row(5, wind_data.usable_hourly_speed)
base_excel.populate_standard_row(6, wind_data.usable_hourly_direction)
base_excel.populate_cloud_row(7, cloud_data.usable_hourly)
base_excel.populate_sun_moon_row(8, sun_moon_data.readable_sun_hourly)
base_excel.populate_sun_moon_row(9, sun_moon_data.readable_moon_hourly)
base_excel.populate_borders()
base_excel.color_scale(1, temperature_data.usable_hourly)
base_excel.color_scale(2, dew_data.usable_hourly_dew_point)
base_excel.precipitation_strength_fill(3, precipitation_data.usable_hourly_strength)
base_excel.precipitation_type_fill(4, precipitation_data.usable_hourly_type)
base_excel.color_scale(5, wind_data.usable_hourly_speed)
base_excel.wind_direction_fill(6)
base_excel.cloud_fill(7, cloud_data.usable_hourly)
base_excel.sun_fill(8, sun_moon_data.readable_sun_hourly)
base_excel.moon_fill(9, sun_moon_data.readable_moon_hourly)

#### Save workbook
base_excel.save_workbook()